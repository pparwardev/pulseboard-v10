"""OT/Leave Tracker API."""
from io import BytesIO
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, extract
from pydantic import BaseModel
from datetime import date, datetime
from typing import Optional
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from app.core.database import get_db
from app.core.security import get_current_user
from app.models.user import User
from app.models.ot_submission import OTSubmission
from app.models.leave_submission import LeaveSubmission
from app.models.upcoming_leave import UpcomingLeave
from app.utils.notification_helper import create_notification
from app.utils.email_helper import send_leave_announce_email
from app.api.profile import photo_url

router = APIRouter(prefix="/api/ot", tags=["OT/Leave"])


@router.get("/monthly-ot-check")
def check_monthly_ot(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Check if user needs to submit OT for previous month. Shows on 1st of each month."""
    today = date.today()
    if today.day > 7:
        return {"show_prompt": False}
    prev_month = today.month - 1 if today.month > 1 else 12
    prev_year = today.year if today.month > 1 else today.year - 1
    has_ot = db.query(OTSubmission).filter(
        OTSubmission.user_id == current_user.id,
        extract('month', OTSubmission.date) == prev_month,
        extract('year', OTSubmission.date) == prev_year
    ).first()
    dismissed_key = f"ot_dismissed_{prev_year}_{prev_month}"
    from app.models.notification import Notification
    dismissed = db.query(Notification).filter(
        Notification.user_id == current_user.id,
        Notification.type == dismissed_key
    ).first()
    if has_ot or dismissed:
        return {"show_prompt": False}
    month_names = ['','January','February','March','April','May','June','July','August','September','October','November','December']
    return {"show_prompt": True, "month": month_names[prev_month], "year": prev_year}


@router.post("/no-ot-this-month")
def no_ot_this_month(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Mark that user had no OT for previous month."""
    today = date.today()
    prev_month = today.month - 1 if today.month > 1 else 12
    prev_year = today.year if today.month > 1 else today.year - 1
    dismissed_key = f"ot_dismissed_{prev_year}_{prev_month}"
    from app.models.notification import Notification
    existing = db.query(Notification).filter(
        Notification.user_id == current_user.id,
        Notification.type == dismissed_key
    ).first()
    if not existing:
        n = Notification(user_id=current_user.id, title="No OT", message=f"No OT for {prev_month}/{prev_year}", type=dismissed_key)
        db.add(n)
        db.commit()
    return {"message": "Noted. No OT for this month."}


class OTSubmit(BaseModel):
    date: str
    start_time: str
    end_time: str
    hours: float
    ot_type: str = "OT"
    reason: Optional[str] = None


class LeaveSubmit(BaseModel):
    leave_date: str
    end_date: Optional[str] = None
    leave_type: str = "PL"
    reason: Optional[str] = None


@router.post("/submit")
def submit_ot(data: OTSubmit, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    ot = OTSubmission(
        user_id=current_user.id, date=data.date, start_time=data.start_time,
        end_time=data.end_time, hours=data.hours, ot_type=data.ot_type, reason=data.reason,
    )
    db.add(ot)
    db.commit()
    return {"message": "OT submitted"}


@router.get("/my-submissions")
def get_my_ot(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    submissions = db.query(OTSubmission).filter(OTSubmission.user_id == current_user.id).order_by(OTSubmission.date.desc()).all()
    return [{"id": s.id, "date": s.date.isoformat(), "start_time": s.start_time, "end_time": s.end_time,
             "hours": s.hours, "ot_type": s.ot_type, "reason": s.reason, "status": s.status,
             "created_at": s.created_at.isoformat() if s.created_at else None} for s in submissions]


@router.post("/leave/submit")
def submit_leave(data: LeaveSubmit, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    leave = LeaveSubmission(user_id=current_user.id, leave_date=data.leave_date, leave_type=data.leave_type, reason=data.reason)
    db.add(leave)
    upcoming = UpcomingLeave(user_id=current_user.id, leave_date=data.leave_date, end_date=data.end_date, leave_type=data.leave_type, reason=data.reason)
    db.add(upcoming)
    db.commit()
    return {"message": "Leave submitted"}


@router.get("/my-leaves")
def get_my_leaves(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    leaves = db.query(LeaveSubmission).filter(LeaveSubmission.user_id == current_user.id).order_by(LeaveSubmission.leave_date.desc()).all()
    return [{"id": l.id, "leave_date": l.leave_date.isoformat(), "leave_type": l.leave_type, "reason": l.reason,
             "created_at": l.created_at.isoformat() if l.created_at else None} for l in leaves]


@router.get("/my-submissions-full")
def get_my_ot_full(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    submissions = db.query(OTSubmission).filter(OTSubmission.user_id == current_user.id).order_by(OTSubmission.date.desc()).all()
    return [{"id": s.id, "date": s.date.isoformat(), "start_time": s.start_time, "end_time": s.end_time,
             "hours": s.hours, "ot_type": s.ot_type, "reason": s.reason, "status": s.status,
             "is_delivered": s.is_delivered, "is_viewed": s.is_viewed,
             "user_name": current_user.name, "login": current_user.login,
             "employee_id": current_user.employee_id or "N/A",
             "created_at": s.created_at.isoformat() if s.created_at else None} for s in submissions]


@router.put("/update/{ot_id}")
def update_ot(ot_id: int, data: OTSubmit, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    ot = db.query(OTSubmission).filter(OTSubmission.id == ot_id, OTSubmission.user_id == current_user.id).first()
    if not ot: raise HTTPException(status_code=404, detail="OT not found")
    if ot.is_delivered: raise HTTPException(status_code=400, detail="Cannot edit delivered OT")
    ot.date = data.date; ot.start_time = data.start_time; ot.end_time = data.end_time
    ot.hours = data.hours; ot.ot_type = data.ot_type; ot.reason = data.reason
    db.commit()
    return {"message": "OT updated"}


@router.delete("/delete/{ot_id}")
def delete_ot(ot_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    ot = db.query(OTSubmission).filter(OTSubmission.id == ot_id, OTSubmission.user_id == current_user.id).first()
    if not ot: raise HTTPException(status_code=404, detail="OT not found")
    if ot.is_delivered: raise HTTPException(status_code=400, detail="Cannot delete delivered OT")
    db.delete(ot)
    db.commit()
    return {"message": "OT deleted"}


@router.get("/monthly-reports")
def get_monthly_ot_reports(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    if current_user.role != "manager":
        raise HTTPException(status_code=403, detail="Only managers can view reports")
    team_ids = _get_team_ids(db, current_user)
    submissions = db.query(OTSubmission, User).join(User, OTSubmission.user_id == User.id).filter(
        OTSubmission.user_id.in_(team_ids)
    ).order_by(OTSubmission.user_id, OTSubmission.date.desc()).all()
    reports = {}
    for s, user in submissions:
        month_key = s.date.strftime("%Y-%m")
        key = f"{s.user_id}_{month_key}"
        if key not in reports:
            reports[key] = {"user_id": s.user_id, "user_name": user.name, "user_login": user.login,
                           "month": month_key, "total_hours": 0, "entries": [],
                           "is_delivered": False, "is_viewed": False, "is_downloaded": False}
        reports[key]["total_hours"] += s.hours
        reports[key]["entries"].append({"id": s.id, "date": s.date.isoformat(), "start_time": s.start_time,
            "end_time": s.end_time, "hours": s.hours, "ot_type": s.ot_type, "status": s.status})
        if s.is_delivered: reports[key]["is_delivered"] = True
        if s.is_viewed: reports[key]["is_viewed"] = True
    return list(reports.values())


@router.post("/mark-viewed/{user_id}/{month}")
def mark_ot_viewed(user_id: int, month: str, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    if current_user.role != "manager":
        raise HTTPException(status_code=403, detail="Access denied")
    try:
        year, month_num = map(int, month.split("-"))
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid month format")
    db.query(OTSubmission).filter(
        OTSubmission.user_id == user_id,
        extract('year', OTSubmission.date) == year,
        extract('month', OTSubmission.date) == month_num
    ).update({"is_viewed": True})
    db.commit()
    return {"message": "Marked as viewed"}


@router.get("/download-all-reports")
def download_all_ot_reports(month: str = None, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    import zipfile
    if current_user.role != "manager":
        raise HTTPException(status_code=403, detail="Access denied")
    team_ids = _get_team_ids(db, current_user)
    query = db.query(OTSubmission, User).join(User, OTSubmission.user_id == User.id).filter(OTSubmission.user_id.in_(team_ids))
    if month:
        try:
            y, m = map(int, month.split("-"))
            query = query.filter(extract('year', OTSubmission.date) == y, extract('month', OTSubmission.date) == m)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid month format")
    submissions = query.order_by(OTSubmission.user_id, OTSubmission.date).all()
    if not submissions:
        raise HTTPException(status_code=404, detail="No OT records found")
    grouped = {}
    for s, user in submissions:
        mk = s.date.strftime("%Y-%m")
        key = f"{s.user_id}_{mk}"
        if key not in grouped: grouped[key] = {"user": user, "month": mk, "subs": []}
        grouped[key]["subs"].append(s)
    zip_buffer = BytesIO()
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for key, data in grouped.items():
            user = data["user"]; ms = data["month"]; subs = data["subs"]
            wb = openpyxl.Workbook(); wb.remove(wb.active)
            for ot_type in ["OT", "NSA", "ASA"]:
                ws = wb.create_sheet(title=ot_type)
                hdrs = ["Employee ID", "Name", "Login", "Date",
                        f"Start Time of the {'OT' if ot_type == 'OT' else 'Shift'}",
                        f"End Time of the {'OT' if ot_type == 'OT' else 'Shift'}", "Number of Hours"]
                for col, h in enumerate(hdrs, 1):
                    cell = ws.cell(row=1, column=col, value=h)
                    cell.fill = header_fill; cell.font = header_font; cell.alignment = Alignment(horizontal="center")
                ts = [s for s in subs if s.ot_type == ot_type]; th = 0
                for row, s in enumerate(ts, 2):
                    ws.cell(row=row, column=1, value=user.employee_id or "N/A")
                    ws.cell(row=row, column=2, value=user.name)
                    ws.cell(row=row, column=3, value=user.login)
                    ws.cell(row=row, column=4, value=s.date.strftime("%d-%m-%Y"))
                    ws.cell(row=row, column=5, value=s.start_time)
                    ws.cell(row=row, column=6, value=s.end_time)
                    ws.cell(row=row, column=7, value=s.hours); th += s.hours
                if ts:
                    tr = len(ts) + 2
                    ws.cell(row=tr, column=6, value="TOTAL").font = Font(bold=True)
                    ws.cell(row=tr, column=7, value=th).font = Font(bold=True)
                for c, w in zip('ABCDEFG', [12, 20, 15, 15, 20, 20, 15]): ws.column_dimensions[c].width = w
            eb = BytesIO(); wb.save(eb); eb.seek(0)
            zf.writestr(f"{user.login}_{ms}_OT.xlsx", eb.getvalue())
    zip_buffer.seek(0)
    return StreamingResponse(zip_buffer, media_type="application/zip",
                             headers={"Content-Disposition": f"attachment; filename=OT_Reports_{month or 'All'}.zip"})


@router.post("/send-to-manager/{user_id}/{month}")
def send_to_manager(user_id: int, month: str, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    if current_user.id != user_id:
        raise HTTPException(status_code=403, detail="Access denied")
    try:
        year, month_num = map(int, month.split("-"))
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid month format")
    db.query(OTSubmission).filter(
        OTSubmission.user_id == user_id,
        extract('year', OTSubmission.date) == year,
        extract('month', OTSubmission.date) == month_num
    ).update({"is_delivered": True})
    db.commit()
    manager = db.query(User).filter(User.team_name == current_user.team_name, User.role == 'manager').first()
    if manager:
        create_notification(db=db, user_id=manager.id, title='New OT Report Submitted',
                            message=f'{current_user.name} has submitted OT report for {month}', notification_type='ot_leave_submitted')
    create_notification(db=db, user_id=current_user.id, title='OT Report Delivered',
                        message=f'Your OT report for {month} has been delivered to manager', notification_type='ot_leave_submitted')
    return {"message": "Report sent to manager successfully"}


@router.get("/download-report/{user_id}/{month}")
def download_ot_report(user_id: int, month: str, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    try:
        year, month_num = map(int, month.split("-"))
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid month format")
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    submissions = db.query(OTSubmission).filter(
        OTSubmission.user_id == user_id,
        extract('year', OTSubmission.date) == year,
        extract('month', OTSubmission.date) == month_num
    ).order_by(OTSubmission.date).all()
    if not submissions:
        raise HTTPException(status_code=404, detail="No OT records found")
    db.query(OTSubmission).filter(
        OTSubmission.user_id == user_id,
        extract('year', OTSubmission.date) == year,
        extract('month', OTSubmission.date) == month_num
    ).update({"is_viewed": True})
    db.commit()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    login = user.login or user.name
    employee_id = getattr(user, 'employee_id', None) or "N/A"
    for ot_type in ["OT", "NSA", "ASA"]:
        ws = wb.create_sheet(title=ot_type)
        headers = ["Employee ID", "Name", "Login", "Date",
                   f"Start Time of the {'OT' if ot_type == 'OT' else 'Shift'}",
                   f"End Time of the {'OT' if ot_type == 'OT' else 'Shift'}", "Number of Hours"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        type_subs = [s for s in submissions if s.ot_type == ot_type]
        total_hours = 0
        for row, s in enumerate(type_subs, 2):
            ws.cell(row=row, column=1, value=employee_id)
            ws.cell(row=row, column=2, value=user.name)
            ws.cell(row=row, column=3, value=login)
            ws.cell(row=row, column=4, value=s.date.strftime("%d-%m-%Y"))
            ws.cell(row=row, column=5, value=s.start_time)
            ws.cell(row=row, column=6, value=s.end_time)
            ws.cell(row=row, column=7, value=s.hours)
            total_hours += s.hours
        if type_subs:
            total_row = len(type_subs) + 2
            ws.cell(row=total_row, column=6, value="TOTAL").font = Font(bold=True)
            ws.cell(row=total_row, column=7, value=total_hours).font = Font(bold=True)
        for c, w in zip('ABCDEFG', [12, 20, 15, 15, 20, 20, 15]):
            ws.column_dimensions[c].width = w
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"{login}_{month}_OT.xlsx"
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})





# ── OTTO (Upcoming Leaves) ──

class UpcomingLeaveCreate(BaseModel):
    leave_date: str
    end_date: Optional[str] = None
    leave_type: str = "sick"
    reason: Optional[str] = None


@router.get("/upcoming-leaves")
def get_my_upcoming_leaves(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    leaves = db.query(UpcomingLeave).filter(UpcomingLeave.user_id == current_user.id).order_by(UpcomingLeave.leave_date.desc()).all()
    return [{"id": l.id, "leave_date": l.leave_date.isoformat(), "end_date": l.end_date.isoformat() if l.end_date else None,
             "leave_type": l.leave_type, "reason": l.reason, "is_announced": l.is_announced} for l in leaves]


@router.post("/upcoming-leaves")
def create_upcoming_leave(data: UpcomingLeaveCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    start = date.fromisoformat(data.leave_date)
    end = date.fromisoformat(data.end_date) if data.end_date else start
    existing = db.query(UpcomingLeave).filter(UpcomingLeave.user_id == current_user.id).all()
    for ex in existing:
        ex_start = ex.leave_date
        ex_end = ex.end_date or ex.leave_date
        if ex_start <= end and ex_end >= start:
            raise HTTPException(status_code=400, detail="You already have leave applied for one or more of these dates")
    leave = UpcomingLeave(user_id=current_user.id, leave_date=start, end_date=end, leave_type=data.leave_type, reason=data.reason)
    db.add(leave)
    db.commit()
    db.refresh(leave)
    return {"id": leave.id, "message": "Leave added"}


@router.put("/upcoming-leaves/{leave_id}")
def update_upcoming_leave(leave_id: int, data: UpcomingLeaveCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    leave = db.query(UpcomingLeave).filter(UpcomingLeave.id == leave_id, UpcomingLeave.user_id == current_user.id).first()
    if not leave: raise HTTPException(status_code=404, detail="Not found")
    leave.leave_date = data.leave_date
    leave.end_date = data.end_date
    leave.leave_type = data.leave_type
    leave.reason = data.reason
    leave.is_announced = True
    db.commit()
    return {"message": "Updated"}


@router.delete("/upcoming-leaves/{leave_id}")
def delete_upcoming_leave(leave_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    leave = db.query(UpcomingLeave).filter(UpcomingLeave.id == leave_id, UpcomingLeave.user_id == current_user.id).first()
    if not leave: raise HTTPException(status_code=404, detail="Not found")
    if leave.is_announced: raise HTTPException(status_code=400, detail="Announced leave cannot be deleted. Use cancel instead.")
    db.delete(leave)
    db.commit()
    return {"message": "Deleted"}


@router.post("/upcoming-leaves/{leave_id}/announce")
def announce_leave(leave_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    leave = db.query(UpcomingLeave).filter(UpcomingLeave.id == leave_id, UpcomingLeave.user_id == current_user.id).first()
    if not leave: raise HTTPException(status_code=404, detail="Not found")
    leave.is_announced = True
    db.commit()
    # Notify manager
    manager = db.query(User).filter(User.team_name == current_user.team_name, User.role == 'manager').first()
    date_str = leave.leave_date.strftime('%d %b %Y')
    end_str = f" to {leave.end_date.strftime('%d %b %Y')}" if leave.end_date and leave.end_date != leave.leave_date else ""
    msg = f"{current_user.name} has announced {leave.leave_type} leave on {date_str}{end_str}"
    if manager:
        create_notification(db=db, user_id=manager.id, title='Leave Announced', message=msg, notification_type='leave_announced')
    # Notify all team members
    team_ids = _get_team_ids(db, current_user)
    for tid in team_ids:
        if tid != current_user.id and (not manager or tid != manager.id):
            create_notification(db=db, user_id=tid, title='Team Leave Update', message=msg, notification_type='leave_announced')
    from datetime import timedelta
    end_dt = leave.end_date if leave.end_date else leave.leave_date
    return_dt = end_dt + timedelta(days=1)
    send_leave_announce_email(
        sender_email=current_user.email,
        login_id=current_user.login,
        user_name=current_user.name,
        leave_start=leave.leave_date.strftime('%d %b %Y'),
        leave_end=end_dt.strftime('%d %b %Y'),
        return_date=return_dt.strftime('%d %b %Y'),
    )
    return {"message": "Leave announced to team & manager"}


@router.post("/upcoming-leaves/{leave_id}/cancel")
def cancel_leave(leave_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    leave = db.query(UpcomingLeave).filter(UpcomingLeave.id == leave_id, UpcomingLeave.user_id == current_user.id).first()
    if not leave: raise HTTPException(status_code=404, detail="Not found")
    date_str = leave.leave_date.strftime('%d %b %Y')
    end_str = f" to {leave.end_date.strftime('%d %b %Y')}" if leave.end_date and leave.end_date != leave.leave_date else ""
    msg = f"{current_user.name} has cancelled {leave.leave_type} leave on {date_str}{end_str}"
    manager = db.query(User).filter(User.team_name == current_user.team_name, User.role == 'manager').first()
    if manager:
        create_notification(db=db, user_id=manager.id, title='Leave Cancelled', message=msg, notification_type='leave_announced')
    team_ids = _get_team_ids(db, current_user)
    for tid in team_ids:
        if tid != current_user.id and (not manager or tid != manager.id):
            create_notification(db=db, user_id=tid, title='Leave Cancelled', message=msg, notification_type='leave_announced')
    db.delete(leave)
    db.commit()
    return {"message": "Leave cancelled and team notified"}


@router.get("/upcoming-leaves/team-announced")
def get_team_announced_leaves(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    team_ids = _get_team_ids(db, current_user)
    leaves = db.query(UpcomingLeave).filter(
        UpcomingLeave.user_id.in_(team_ids), UpcomingLeave.is_announced == True
    ).order_by(UpcomingLeave.leave_date.desc()).all()
    result = []
    for l in leaves:
        user = db.query(User).filter(User.id == l.user_id).first()
        result.append({
            "id": l.id, "user_id": l.user_id, "user_name": user.name if user else "Unknown",
            "employee_id": user.login if user else None,
            "profile_picture": photo_url(user),
            "leave_date": l.leave_date.isoformat(), "end_date": l.end_date.isoformat() if l.end_date else l.leave_date.isoformat(),
            "leave_type": l.leave_type, "reason": l.reason,
        })
    return result


@router.get("/ot-champions")
def get_ot_champions(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Last month OT leaderboard for the team."""
    today = date.today()
    if today.month == 1:
        lm_year, lm_month = today.year - 1, 12
    else:
        lm_year, lm_month = today.year, today.month - 1
    team_ids = _get_team_ids(db, current_user)
    rows = db.query(
        OTSubmission.user_id, func.sum(OTSubmission.hours).label("total_hours")
    ).filter(
        OTSubmission.user_id.in_(team_ids),
        extract('year', OTSubmission.date) == lm_year,
        extract('month', OTSubmission.date) == lm_month,
    ).group_by(OTSubmission.user_id).order_by(func.sum(OTSubmission.hours).desc()).all()
    result = []
    for uid, total in rows:
        u = db.query(User).filter(User.id == uid).first()
        if u:
            result.append({
                "user_id": uid, "name": u.name, "login": u.login,
                "profile_picture": photo_url(u),
                "total_hours": round(total, 1),
            })
    from calendar import month_name
    return {"month": f"{month_name[lm_month]} {lm_year}", "leaderboard": result}


def _get_team_ids(db, current_user):
    if current_user.role == 'manager':
        ids = [uid for (uid,) in db.query(User.id).filter(User.manager_login == current_user.login, User.is_active == True).all()]
    else:
        ids = [uid for (uid,) in db.query(User.id).filter(User.manager_login == current_user.manager_login, User.is_active == True).all()]
    if current_user.id not in ids:
        ids.append(current_user.id)
    return ids
