"""
OT Submit Module - FastAPI Router (Embeddable)

All endpoints for OT submission: CRUD, send-to-manager, download report,
monthly reports (manager), download-all (manager), OT champions leaderboard.

Usage:
    from ot_router import create_ot_router

    router = create_ot_router(
        get_db=get_db,
        get_current_user=get_current_user,
        get_team_ids=lambda db, user: [...],
        User=User,
        create_notification=my_notify_fn,  # optional
    )
    app.include_router(router, prefix="/api/ot")
"""
from io import BytesIO
from datetime import date
from typing import Callable, Optional, List
from calendar import month_name

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import func, extract

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from .ot_model import OTSubmission


class OTSubmitPayload(BaseModel):
    date: str
    start_time: str
    end_time: str
    hours: float
    ot_type: str = "OT"
    reason: Optional[str] = None


def create_ot_router(
    get_db: Callable,
    get_current_user: Callable,
    get_team_ids: Callable,          # (db, user) -> List[int]
    User,                            # Your User SQLAlchemy model
    create_notification: Optional[Callable] = None,  # (db, user_id, title, message, notification_type)
) -> APIRouter:
    router = APIRouter(tags=["OT Submit"])

    # ── CRUD ────────────────────────────────────────────────────────────

    @router.post("/submit")
    def submit_ot(data: OTSubmitPayload, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
        ot = OTSubmission(
            user_id=current_user.id, date=data.date, start_time=data.start_time,
            end_time=data.end_time, hours=data.hours, ot_type=data.ot_type, reason=data.reason,
        )
        db.add(ot)
        db.commit()
        return {"message": "OT submitted"}

    @router.get("/my-submissions")
    def get_my_ot(db: Session = Depends(get_db), current_user=Depends(get_current_user)):
        subs = db.query(OTSubmission).filter(OTSubmission.user_id == current_user.id).order_by(OTSubmission.date.desc()).all()
        return [{"id": s.id, "date": s.date.isoformat(), "start_time": s.start_time, "end_time": s.end_time,
                 "hours": s.hours, "ot_type": s.ot_type, "reason": s.reason, "status": s.status,
                 "created_at": s.created_at.isoformat() if s.created_at else None} for s in subs]

    @router.get("/my-submissions-full")
    def get_my_ot_full(db: Session = Depends(get_db), current_user=Depends(get_current_user)):
        subs = db.query(OTSubmission).filter(OTSubmission.user_id == current_user.id).order_by(OTSubmission.date.desc()).all()
        return [{"id": s.id, "date": s.date.isoformat(), "start_time": s.start_time, "end_time": s.end_time,
                 "hours": s.hours, "ot_type": s.ot_type, "reason": s.reason, "status": s.status,
                 "is_delivered": s.is_delivered, "is_viewed": s.is_viewed,
                 "user_name": current_user.name, "login": getattr(current_user, 'login', None),
                 "employee_id": getattr(current_user, 'employee_id', None) or "N/A",
                 "created_at": s.created_at.isoformat() if s.created_at else None} for s in subs]

    @router.put("/update/{ot_id}")
    def update_ot(ot_id: int, data: OTSubmitPayload, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
        ot = db.query(OTSubmission).filter(OTSubmission.id == ot_id, OTSubmission.user_id == current_user.id).first()
        if not ot:
            raise HTTPException(404, "OT not found")
        if ot.is_delivered:
            raise HTTPException(400, "Cannot edit delivered OT")
        ot.date = data.date
        ot.start_time = data.start_time
        ot.end_time = data.end_time
        ot.hours = data.hours
        ot.ot_type = data.ot_type
        ot.reason = data.reason
        db.commit()
        return {"message": "OT updated"}

    @router.delete("/delete/{ot_id}")
    def delete_ot(ot_id: int, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
        ot = db.query(OTSubmission).filter(OTSubmission.id == ot_id, OTSubmission.user_id == current_user.id).first()
        if not ot:
            raise HTTPException(404, "OT not found")
        if ot.is_delivered:
            raise HTTPException(400, "Cannot delete delivered OT")
        db.delete(ot)
        db.commit()
        return {"message": "OT deleted"}

    # ── Send to Manager ─────────────────────────────────────────────────

    @router.post("/send-to-manager/{user_id}/{month}")
    def send_to_manager(user_id: int, month: str, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
        if current_user.id != user_id:
            raise HTTPException(403, "Access denied")
        try:
            year, month_num = map(int, month.split("-"))
        except ValueError:
            raise HTTPException(400, "Invalid month format")
        db.query(OTSubmission).filter(
            OTSubmission.user_id == user_id,
            extract('year', OTSubmission.date) == year,
            extract('month', OTSubmission.date) == month_num
        ).update({"is_delivered": True})
        db.commit()
        if create_notification:
            manager = db.query(User).filter(User.team_name == current_user.team_name, User.role == 'manager').first()
            if manager:
                create_notification(db=db, user_id=manager.id, title='New OT Report Submitted',
                                    message=f'{current_user.name} has submitted OT report for {month}',
                                    notification_type='ot_leave_submitted')
            create_notification(db=db, user_id=current_user.id, title='OT Report Delivered',
                                message=f'Your OT report for {month} has been delivered to manager',
                                notification_type='ot_leave_submitted')
        return {"message": "Report sent to manager successfully"}

    # ── Download Report (Individual) ────────────────────────────────────

    @router.get("/download-report/{user_id}/{month}")
    def download_ot_report(user_id: int, month: str, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
        try:
            year, month_num = map(int, month.split("-"))
        except ValueError:
            raise HTTPException(400, "Invalid month format")
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            raise HTTPException(404, "User not found")
        submissions = db.query(OTSubmission).filter(
            OTSubmission.user_id == user_id,
            extract('year', OTSubmission.date) == year,
            extract('month', OTSubmission.date) == month_num
        ).order_by(OTSubmission.date).all()
        if not submissions:
            raise HTTPException(404, "No OT records found")
        # Mark as viewed
        db.query(OTSubmission).filter(
            OTSubmission.user_id == user_id,
            extract('year', OTSubmission.date) == year,
            extract('month', OTSubmission.date) == month_num
        ).update({"is_viewed": True})
        db.commit()
        wb = _build_ot_workbook(user, submissions)
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        login = getattr(user, 'login', None) or user.name
        return StreamingResponse(output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={login}_{month}_OT.xlsx"})

    # ── Manager: Monthly Reports ────────────────────────────────────────

    @router.get("/monthly-reports")
    def get_monthly_ot_reports(db: Session = Depends(get_db), current_user=Depends(get_current_user)):
        if current_user.role != "manager":
            raise HTTPException(403, "Only managers can view reports")
        team_ids = get_team_ids(db, current_user)
        submissions = db.query(OTSubmission, User).join(User, OTSubmission.user_id == User.id).filter(
            OTSubmission.user_id.in_(team_ids)
        ).order_by(OTSubmission.user_id, OTSubmission.date.desc()).all()
        reports = {}
        for s, user in submissions:
            mk = s.date.strftime("%Y-%m")
            key = f"{s.user_id}_{mk}"
            if key not in reports:
                reports[key] = {"user_id": s.user_id, "user_name": user.name,
                                "user_login": getattr(user, 'login', None),
                                "month": mk, "total_hours": 0, "entries": [],
                                "is_delivered": False, "is_viewed": False}
            reports[key]["total_hours"] += s.hours
            reports[key]["entries"].append({"id": s.id, "date": s.date.isoformat(), "start_time": s.start_time,
                "end_time": s.end_time, "hours": s.hours, "ot_type": s.ot_type, "status": s.status})
            if s.is_delivered:
                reports[key]["is_delivered"] = True
            if s.is_viewed:
                reports[key]["is_viewed"] = True
        return list(reports.values())

    @router.post("/mark-viewed/{user_id}/{month}")
    def mark_ot_viewed(user_id: int, month: str, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
        if current_user.role != "manager":
            raise HTTPException(403, "Access denied")
        try:
            year, month_num = map(int, month.split("-"))
        except ValueError:
            raise HTTPException(400, "Invalid month format")
        db.query(OTSubmission).filter(
            OTSubmission.user_id == user_id,
            extract('year', OTSubmission.date) == year,
            extract('month', OTSubmission.date) == month_num
        ).update({"is_viewed": True})
        db.commit()
        return {"message": "Marked as viewed"}

    # ── Manager: Download All Reports (ZIP) ─────────────────────────────

    @router.get("/download-all-reports")
    def download_all_ot_reports(month: str = None, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
        import zipfile
        if current_user.role != "manager":
            raise HTTPException(403, "Access denied")
        team_ids = get_team_ids(db, current_user)
        query = db.query(OTSubmission, User).join(User, OTSubmission.user_id == User.id).filter(
            OTSubmission.user_id.in_(team_ids))
        if month:
            try:
                y, m = map(int, month.split("-"))
                query = query.filter(extract('year', OTSubmission.date) == y, extract('month', OTSubmission.date) == m)
            except ValueError:
                raise HTTPException(400, "Invalid month format")
        submissions = query.order_by(OTSubmission.user_id, OTSubmission.date).all()
        if not submissions:
            raise HTTPException(404, "No OT records found")
        grouped = {}
        for s, user in submissions:
            mk = s.date.strftime("%Y-%m")
            key = f"{s.user_id}_{mk}"
            if key not in grouped:
                grouped[key] = {"user": user, "month": mk, "subs": []}
            grouped[key]["subs"].append(s)
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            for key, data in grouped.items():
                user = data["user"]
                wb = _build_ot_workbook(user, data["subs"])
                eb = BytesIO()
                wb.save(eb)
                eb.seek(0)
                login = getattr(user, 'login', None) or user.name
                zf.writestr(f"{login}_{data['month']}_OT.xlsx", eb.getvalue())
        zip_buffer.seek(0)
        return StreamingResponse(zip_buffer, media_type="application/zip",
                                 headers={"Content-Disposition": f"attachment; filename=OT_Reports_{month or 'All'}.zip"})

    # ── OT Champions Leaderboard ────────────────────────────────────────

    @router.get("/ot-champions")
    def get_ot_champions(db: Session = Depends(get_db), current_user=Depends(get_current_user)):
        today = date.today()
        lm_year, lm_month = (today.year - 1, 12) if today.month == 1 else (today.year, today.month - 1)
        team_ids = get_team_ids(db, current_user)
        rows = db.query(
            OTSubmission.user_id, func.sum(OTSubmission.hours).label("total_hours")
        ).filter(
            OTSubmission.user_id.in_(team_ids),
            extract('year', OTSubmission.date) == lm_year,
            extract('month', OTSubmission.date) == lm_month,
        ).group_by(OTSubmission.user_id).order_by(func.sum(OTSubmission.hours).desc()).all()
        result = []
        for uid, total in rows:
            u = db.query(User).filter(User.id == uid).first()
            if u:
                result.append({
                    "user_id": uid, "name": u.name,
                    "login": getattr(u, 'login', None),
                    "profile_picture": getattr(u, 'profile_picture', None),
                    "total_hours": round(total, 1),
                })
        return {"month": f"{month_name[lm_month]} {lm_year}", "leaderboard": result}

    return router


# ── Helper: Build Excel Workbook ────────────────────────────────────────

def _build_ot_workbook(user, submissions):
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    login = getattr(user, 'login', None) or user.name
    employee_id = getattr(user, 'employee_id', None) or "N/A"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for ot_type in ["OT", "NSA", "ASA"]:
        ws = wb.create_sheet(title=ot_type)
        headers = ["Employee ID", "Name", "Login", "Date",
                   f"Start Time of the {'OT' if ot_type == 'OT' else 'Shift'}",
                   f"End Time of the {'OT' if ot_type == 'OT' else 'Shift'}", "Number of Hours"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        type_subs = [s for s in submissions if s.ot_type == ot_type]
        total_hours = 0
        for row, s in enumerate(type_subs, 2):
            ws.cell(row=row, column=1, value=employee_id)
            ws.cell(row=row, column=2, value=user.name)
            ws.cell(row=row, column=3, value=login)
            ws.cell(row=row, column=4, value=s.date.strftime("%d-%m-%Y"))
            ws.cell(row=row, column=5, value=s.start_time)
            ws.cell(row=row, column=6, value=s.end_time)
            ws.cell(row=row, column=7, value=s.hours)
            total_hours += s.hours
        if type_subs:
            tr = len(type_subs) + 2
            ws.cell(row=tr, column=6, value="TOTAL").font = Font(bold=True)
            ws.cell(row=tr, column=7, value=total_hours).font = Font(bold=True)
        for c, w in zip('ABCDEFG', [12, 20, 15, 15, 20, 20, 15]):
            ws.column_dimensions[c].width = w
    return wb
